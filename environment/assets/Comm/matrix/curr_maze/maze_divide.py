import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import chardet

collision_id = 20000


def detect_encoding(file_path):
    with open(file_path, "rb") as f:
        raw_data = f.read()
    result = chardet.detect(raw_data)
    return result["encoding"]


# 读取各层次属性文件
maze_path = "environment\\assets\Comm\matrix\\curr_maze"

sector_encoding = detect_encoding(f"{maze_path}\\blocks\\sector_blocks.csv")
sector_df = pd.read_csv(
    f"{maze_path}\\blocks\\sector_blocks.csv",
    header=None,
    names=["ID", "World", "Sector"],
    encoding=sector_encoding,
)
area_encoding = detect_encoding(f"{maze_path}\\blocks\\area_blocks.csv")
area_df = pd.read_csv(
    f"{maze_path}\\blocks\\area_blocks.csv",
    header=None,
    names=["ID", "World", "Sector", "Area"],
    encoding=area_encoding,
)
object_encoding = detect_encoding(f"{maze_path}\\blocks\\object_blocks.csv")
object_df = pd.read_csv(
    f"{maze_path}\\blocks\\object_blocks.csv",
    header=None,
    names=["ID", "World", "All", "Object"],
    encoding=object_encoding,
)

# 读取maze.xlsx中的地图数据
maze_df = pd.read_excel(
    f"{maze_path}\\new_maze.xlsx",
    header=None,
)

maze_wb = load_workbook(
    f"{maze_path}\\new_maze.xlsx",
)
ws = maze_wb.active

# 创建结果DataFrame
sector_maze = maze_df.copy()
area_maze = maze_df.copy()
object_maze = maze_df.copy()
collision_maze = maze_df.copy()


# 定义解析函数
def parse_value(value):
    # 反转层次顺序，获取object, area, sector
    if value == "*":
        return 0, 0, 0
    parts = value.split(" ")
    parts.reverse()
    # print("--- ", parts)

    if len(parts) == 1:
        # 只有一个层次，即只存在sector层次
        return sector_df[sector_df["Sector"] == parts[0]]["ID"].values[0], 0, 0

    elif len(parts) == 2:
        # 存在area和sector
        sector_desc = parts[0]
        area_desc = parts[1]

        # 获取sector ID
        sector_id = sector_df[sector_df["Sector"] == sector_desc]["ID"].values[0]

        # 获取area ID
        area_id = area_df[
            (area_df["Sector"] == sector_desc) & (area_df["Area"] == area_desc)
        ]["ID"].values[0]

        return sector_id, area_id, 0

    elif len(parts) == 3:
        # print(parts)
        # 存在object, area和sector
        sector_desc = parts[0]
        area_desc = parts[1]
        object_desc = parts[2]

        # 获取sector ID
        sector_id = sector_df[sector_df["Sector"] == sector_desc]["ID"].values[0]

        # 获取area ID
        area_id = area_df[
            (area_df["Sector"] == sector_desc) & (area_df["Area"] == area_desc)
        ]["ID"].values[0]

        # 获取object ID
        try:
            object_id = object_df[(object_df["Object"] == object_desc)]["ID"].values[0]
        except:
            print(object_desc, "not in object_blocks")

        return sector_id, area_id, object_id


# 遍历maze_df的每个单元格，解析并填写ID到相应的DataFrame
for row in range(maze_df.shape[0]):
    for col in range(maze_df.shape[1]):
        cur_cell = maze_df.iloc[row, col]
        # print(type(cur_cell), cur_cell, row, col)

        if pd.isna(cur_cell):
            sector_id, area_id, object_id = 0, 0, 0
        else:
            # 解析value
            sector_id, area_id, object_id = parse_value(cur_cell)

        # 更新结果DataFrame
        sector_maze.iloc[row, col] = sector_id
        area_maze.iloc[row, col] = area_id
        object_maze.iloc[row, col] = object_id

# white_fill = PatternFill("solid", fgColor=["000000", "000000"])

for row in ws.iter_rows():
    for cell in row:
        color = cell.fill.start_color.index
        if color != "00000000":
            collision_maze.iloc[cell.row - 1, cell.column - 1] = collision_id
        else:
            collision_maze.iloc[cell.row - 1, cell.column - 1] = 0

# 将结果保存为CSV文件
sector_maze.to_csv(
    f"{maze_path}\\maze\\sector_maze.csv",
    index=False,
    header=False,
    encoding="utf-8",
)
area_maze.to_csv(
    f"{maze_path}\\maze\\area_maze.csv",
    index=False,
    header=False,
    encoding="utf-8",
)
object_maze.to_csv(
    f"{maze_path}\\maze\\object_maze.csv",
    index=False,
    header=False,
    encoding="utf-8",
)
collision_maze.to_csv(
    f"{maze_path}\\maze\\collision_maze.csv",
    index=False,
    header=False,
    encoding="utf-8",
)

sector_df.to_csv(
    f"{maze_path}\\blocks\\sector_blocks.csv",
    index=False,
    header=False,
    encoding="utf-8",
)
area_df.to_csv(
    f"{maze_path}\\blocks\\area_blocks.csv",
    index=False,
    header=False,
    encoding="utf-8",
)
object_df.to_csv(
    f"{maze_path}\\blocks\\object_blocks.csv",
    index=False,
    header=False,
    encoding="utf-8",
)

print("文件已成功保存！")
